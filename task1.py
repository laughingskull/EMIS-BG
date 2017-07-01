import random

from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill
book = Workbook()

sheet = book.active
sheet.title = "Data"

#styling begins

sheet.row_dimensions[1].fill = PatternFill("solid", fgColor="DAE1F1")
sheet.row_dimensions[1].font = Font(bold=True)
sheet['A1'] = "Name"
sheet['B1'] = "Age"
sheet['C1'] = "Score"
sheet['E1'] = "Average Score"

for cell in sheet['B2:C100002']:
 cell = Alignment(horizontal="right")
 
for r in range(3,100002,2):
 sheet.row_dimensions[r].font = Font(color="839B77")
 
#styling ends

names = ['Ivan','Valentina','George','Ivelina','Peter','Gergana','Christian','Maria','Alexander','Stefka']

sumScores = 0

for R in range(2,100002):
 sheet['A' + str(R)] = names[random.randint(0,9)]
 sheet['B' + str(R)] = random.randint(20,80)
 sheet['C' + str(R)] = random.randint(0,100)
 if R <= 101:
  sumScores += sheet['C' + str(R)].value
 
sheet['E2'] = sumScores / 100
 
book.save("scores.xlsx")